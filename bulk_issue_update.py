import requests
import openpyxl


def rest_PUT_Issue_Edit(ticket_key,pr_head_name, pm_name,prog_man_name):

    json_data = {
    "fields": {
        "customfield_11177": {
            "value": pr_head_name
                },
        "customfield_11178": {
            "value": pm_name
                },
        "customfield_11183": {
            "value": prog_man_name
                }
            }
    }

    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + str(ticket_key)
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA=="  #Gobi's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.put(url=endpoint, headers=header_values,json=json_data, timeout=10)

    if r.status_code == 204:
        returned_response = "Updated"
    else:
        returned_response = "Erron when updating the Issue_" + str(ticket_key) + "_<<RestCallERROR>>_" + str(r.status_code) + "---" + str(r.json())

    return returned_response


if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "C:\\Users\\Roshinie_105953\\Downloads\\New Projects_Update_Execution_3.xlsx"
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 2  # < -- CONFIGURATION
    i_max = 5 # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_key = xl_sheet.cell(row=i, column=1).value  # < -- Config
        exis_ph_name = xl_sheet.cell(row=i, column=2).value  # < -- Config
        exis_pm_name = xl_sheet.cell(row=i, column=3).value  # < -- Config
        exis_pr_man_name = xl_sheet.cell(row=i, column=4).value  # < -- Config

        drop_ph_name = xl_sheet.cell(row=i, column=5).value  # < -- Config
        drop_pm_name = xl_sheet.cell(row=i, column=6).value  # < -- Config
        drop_pr_man_name = xl_sheet.cell(row=i, column=7).value  # < -- Config

        if [ticket_key, drop_ph_name , drop_pm_name , drop_pr_man_name ] is not None:

            update_result = rest_PUT_Issue_Edit(ticket_key,drop_ph_name,drop_pm_name,drop_pr_man_name)
            print (str(ticket_key) + " - - - >> " + str(update_result))
            xl_sheet.cell(row=i, column=8).value = update_result

    xl_workbook.save(path)