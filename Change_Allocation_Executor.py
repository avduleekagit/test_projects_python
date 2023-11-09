
from Child_issue_create import rest_get_Parent_Details,rest_Post_Issue_Create,rest_get_full_jql_result,rest_Post_Issue_Transisiton,rest_Post_Issue_AddComment
from Check_Over_Allocation import rest_get_Allocation_check_v_1_0
from Working_Day_checker import get_working_days
import openpyxl
from datetime import datetime
from check_user_status import rest_get_user_validations_jql_call # For 18/05/23 CR user_max_hours_get

def validate_change_request(change_tkt_obj,orig_tkt_obj):

    # -- Check Project Code Same
    if change_tkt_obj.project_dyn_code != orig_tkt_obj.project_dyn_code:
        return_message ="Error : Project codes are not same !"
        return return_message
   
    # -- Check Change date range is within the original allocation date range
    if (change_tkt_obj.alloc_start == orig_tkt_obj.alloc_start and change_tkt_obj.confirmed_end <= orig_tkt_obj.confirmed_end) or \
         (change_tkt_obj.alloc_start >= orig_tkt_obj.alloc_start and change_tkt_obj.confirmed_end == orig_tkt_obj.confirmed_end):
                return_message ="Proceed"
    else:
        return_message="Error : Change request data range validation failed !"
        return return_message
    
    # -- Over Allocation Check
    # --------------------  18/05/23 CR - get resource max_work_hours to over_allocation_check
    jql = "project = JRT AND issuetype = 'RM Resource Creation' AND status = Active AND 'JRT_Resource Email[Short text]' ~ " + "'" + str(change_tkt_obj.resource_email) + "'"
    req_fields= "customfield_10978"
    get_resrce_card_detail_status_code, get_resrce_card_detail_response_json = rest_get_user_validations_jql_call (jql,req_fields)

    if get_resrce_card_detail_status_code == 200:
             if get_resrce_card_detail_response_json['total'] == 1:
                max_hours_allocation = get_resrce_card_detail_response_json['issues'][0]['fields']['customfield_10978']  
                print ("max_hours_allocation for user is --> " + str(max_hours_allocation))      
             else:
                  return_message = "Error : Getting Max_Working_Hours for this user"
                  return return_message         
    else:
        return_message = "Error REST Calling resource_card_details_check ! " + str(get_resrce_card_detail_response_json)
        return return_message
    
    # -- Logic ---
    if change_tkt_obj.is_over_allocation == "Yes":  ## -- 23/06/2023 Handle Over_allocation = yes scenario
         return_message="Proceed"
         print ("# --------------- Over Allocation Check Skipped ..... !!! ")
    elif change_tkt_obj.hou_per_day <= orig_tkt_obj.hou_per_day:
          return_message="Proceed"
    else:
          req_hours_differ = round(change_tkt_obj.hou_per_day - orig_tkt_obj.hou_per_day,2)
          get_rest_alloc_check= rest_get_Allocation_check_v_1_0(change_tkt_obj.resource_email,change_tkt_obj.alloc_start,change_tkt_obj.confirmed_end, req_hours_differ,max_hours_allocation)
          return_message=get_rest_alloc_check
   
    return return_message


#####################################################

def Process_change_request(change_ticket_obj,orig_ticket_obj):
      response_validate_change_request = validate_change_request(change_ticket_obj,orig_ticket_obj)

      if response_validate_change_request =="Proceed":
            ## Start executions -- - - - -
            jql="project = JRT AND issuetype = 'Daily Allocation' AND status = Allocated AND 'JRT_Demand Reference ID[Short text]' ~ " + str(orig_ticket_obj.key) + " AND 'JRT_Date[Date]' >= " + str(change_ticket_obj.alloc_start) + " AND 'JRT_Date[Date]' <= " + str(change_ticket_obj.confirmed_end)
            req_fields= "key"

            get_child_issues_to_cancel= rest_get_full_jql_result(jql,req_fields)
            print(get_child_issues_to_cancel)
            
            # -- Cancel Issues + Add Comment
            issue_comment="This task was cancelled due to the new Change Allocation - " + str(change_ticket_obj.key)
            returned_response = "Cancelling Tickets --->> "
            for cancel_issue_key in get_child_issues_to_cancel:
                  print("Cancelling . . . " + str(cancel_issue_key))
                  rest_issue_cancel_response = rest_Post_Issue_Transisiton(cancel_issue_key,21)
                  rest_issue_addComment_response = rest_Post_Issue_AddComment(cancel_issue_key,issue_comment)

                  if (rest_issue_cancel_response != "Transitioned") or (rest_issue_addComment_response != "CommentAdded"):
                        returned_response = str(returned_response) + " | " + str(rest_issue_cancel_response) +  " | " + str(rest_issue_addComment_response)
                        return returned_response
                  else:
                        returned_response = str(returned_response) + " | " + str(cancel_issue_key)
            
            #  -- Create New Issues
            returned_response = str(returned_response) + " | " + "Creating Sub Tasks --->> "
            working_days = get_working_days(change_ticket_obj.alloc_start, change_ticket_obj.confirmed_end)
            task_sq = 0
            for item in working_days:
                  sub_task_id = rest_Post_Issue_Create(change_ticket_obj, task_sq, item)
                  terminal_output = str(task_sq) + "-->" + str(item) + "-->" + str(sub_task_id)
                  print(terminal_output)
            
                  returned_response = str(returned_response) + " | " + str(terminal_output)
                  task_sq = task_sq + 1

                  ## ----  Fix for if 403 error fired during executions for 1 or more issues
                  if sub_task_id[0:14] =="Issue Creating":
                        return returned_response        
                  ### -----------------------------------------------------------------------  
            
            returned_response = str(returned_response) + " | _Process_Completed_"  # -- To check whther process completed fully

      else:
            returned_response = response_validate_change_request
      
      return returned_response

#########################################################

if __name__ == "__main__":
   # stuff only to run when not called via 'import' here
    
    # ------- Configuration 
    path = "C:\\Users\\Gobiananth_105358.ADL\\Downloads\\Change_Allocation_List.xlsx" #< -- CONFIGURATION.
    i_start = 2  # < -- CONFIGURATION
    i_max = 3  # < -- CONFIGURATION
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active
    # ------- Configuration 
   
    for i in range(i_start, i_max):
        ticket_id = xl_sheet.cell(row=i, column=1).value  # < -- Config
        ticket_id_status = xl_sheet.cell(row=i, column=2).value  # < -- Config

        if ticket_id is not None: 
            if ticket_id_status != "Processed" :
                get_change_ticket_obj = rest_get_Parent_Details(ticket_id)
                get_orig_ticket_obj = rest_get_Parent_Details(get_change_ticket_obj.orig_tkt_id_for_change)
                response_Process_change_request = Process_change_request(get_change_ticket_obj,get_orig_ticket_obj)
                print(response_Process_change_request)

                xl_sheet.cell(row=i, column=3).value = response_Process_change_request
                xl_sheet.cell(row=i, column=2).value = "Processed"
            else: 
                print ("Record already processed , Hence Skipped")
             
        xl_workbook.save(path)
    
    print("Process Completed" + " ==> " + str(datetime.now()))
      
      
      
  




