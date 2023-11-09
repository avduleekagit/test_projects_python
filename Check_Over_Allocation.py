import requests
import logging
import openpyxl
import json
import math
import logging
import datetime

# Logger configurations
file_name = "Check_Over_Allocation_Log_" + str (datetime.date.strftime(datetime.datetime.now(), "%Y-%m-%d__%H-%M-%S"))
logger_file_path = "D:\\Script_Outputs\\" + str(file_name) + ".txt"
logging.basicConfig(filename=logger_file_path,
                    format='%(asctime)s : %(levelname)s : %(message)s',
                    filemode='a', level=logging.INFO)
logger = logging.getLogger()
##################################################

def rest_get_Allocation_check(inp_email_id,inp_start_date,inp_end_date, inp_req_hours):

    # -->> This is because the user cannot have a single record in system with a allocation of > req_alloc
    #inp_req_hours = 8 - float(inp_req_hours)  #-Bcoz 6.4 usecase failed with 1.5999999999999996
    inp_req_hours = 8 - inp_req_hours
    inp_req_hours = round(inp_req_hours,2)

    #created_jql = "project = JRT AND issuetype = 'Daily Allocation' AND status = 'Allocated' AND 'JRT_Resource Email' ~ '" +  str(inp_email_id) + "' and JRT_Date >= '" + str(inp_start_date)[0:10] + "' and  JRT_Date <= '" + str(inp_end_date)[0:10] + "'"
    created_jql = "project = JRT AND issuetype = 'Daily Allocation' AND status = 'Allocated' AND 'JRT_Resource Email' ~ '" +  str(inp_email_id) + "' and JRT_Date >= '" + str(inp_start_date)[0:10] + "' and  JRT_Date <= '" + str(inp_end_date)[0:10] + "' and 'JRT_Hours per Day[Number]' > " + str(inp_req_hours)
    pagination_startAt = 0 ## -- No Need
    maxResults = 100 ## -- No Need

    #print ("Printing the JQL --->>>>>  " + str(created_jql))
    
    endpoint = "https://aifel.atlassian.net/rest/api/3/search?"
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA==" # Gobi's Tkn

    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    input_params = {'jql': created_jql,'fields': "customfield_11024", 'maxResults' : maxResults, 'startAt': pagination_startAt}

    r = requests.get(url=endpoint, headers=header_values,params=input_params, timeout=10)

    if r.status_code == 200:
        response = r.json()
        if len(response) > 0:
            record_count = response['total']
            if record_count > 0:
                #returned_total_alloc_hours = sum(issue["fields"]["customfield_11024"] for issue in response["issues"])
                returned_total_alloc_status = "One or more excess allocation('s) already been locked for the resource, Hence Rejected --->>> " + str(created_jql)
            else:
                #loop_count = math.ceil(returned_record_count/50)
                returned_total_alloc_status = "Proceed"
                                
    else:
        returned_total_alloc_status  = "_RestCallERROR_ --> " + str(r.status_code)

    return returned_total_alloc_status



###################################################################

def rest_get_Allocation_check_v_1_0(inp_email_id,inp_start_date,inp_end_date, inp_req_hours,inp_max_hours):  ### 18/05/23 CR get max_hours from resource card

    # -->> This is because the user cannot have a single record in system with a allocation of > req_alloc
    #inp_req_hours = 8 - float(inp_req_hours)  #-Bcoz 6.4 usecase failed with 1.5999999999999996
   
    inp_req_hours = inp_max_hours - inp_req_hours ### 18/05/23 CR get max_hours from resource card
    inp_req_hours = round(inp_req_hours,2)
    print ("Allocation maximum available hours required is -->"  + str(inp_req_hours) )

    rest_status_code, rest_payload  = rest_get_Allocation_check_API_Call_Loop(inp_email_id,inp_start_date,inp_end_date,0)

    if rest_status_code == 200:
        record_count = rest_payload['total']
        sums_by_date = {}
        if record_count > 0:
                loop_count = math.ceil( record_count / 100 ) + 1    ## To execute the for loop accordingly ie ( For 1 to 1) stops at 1 itself
                record_no = 0
                logger.info(inp_email_id + " - loop count is " + str(loop_count -1) + " for total records - " + str(record_count) + " ============== ")
                
                for i in range(1, loop_count):
                    logger.info(inp_email_id + " -> Validating balance hours " + str(inp_req_hours) + " : -> -> ->  processing loop " + str(i) + " of " + str(record_count) + " records" + " ---> starting with " + str(record_no))
                    rest_status_code, rest_payload  = rest_get_Allocation_check_API_Call_Loop(inp_email_id,inp_start_date,inp_end_date,record_no)
                    
                    for issue in rest_payload['issues']:
                        date = issue['fields']['customfield_11030']
                        allo_hours = issue['fields']['customfield_11024']
                        #sums_by_date[date] = sums_by_date.get(date, 0) + allo_hours
                        sums_by_date[date] = round(sums_by_date.get(date, 0) + allo_hours, 2)  ## number rounding error fixed
                                       
                    #Individual Loop overallocation check - to bypass rest of the execution through return statement
                    over_allocated_single_loop = {k:v for k, v in sums_by_date.items() if v > inp_req_hours}
                    if len(over_allocated_single_loop) > 0:
                        returned_total_alloc_status = "One or more excess allocation('s) already been locked for the resource, Hence Rejected --->>> " + str(dict(sorted(over_allocated_single_loop.items())))
                        logger.warning (inp_email_id + " - exiting loop in --> loop " + str(i) + " - " + returned_total_alloc_status)
                        return returned_total_alloc_status
                    
                    #Increase record count by 100 for next page in loop
                    record_no = record_no + 100
                
               #Loop complete  : Execute Final Check with all reocords generated by complete loops -> (No need ? Coz relavent checks already done in last loop , only happy path handling is enough here)
                over_allocated_final_check = {k:v for k, v in sums_by_date.items() if v > inp_req_hours}
                if len(over_allocated_final_check) > 0:
                    returned_total_alloc_status = "One or more excess allocation('s) already been locked for the resource, Hence Rejected --->>> " + str(dict(sorted(over_allocated_final_check.items())))
                else:
                    returned_total_alloc_status = "Proceed"

                logger.info (inp_email_id + " ==> " + returned_total_alloc_status)

                # Happy path - just print existing total allocation for given user    
                sorted_sums = dict(sorted(sums_by_date.items())) 
                logger.info(inp_email_id + " - Total allocation list as of now ( total day wise report ) - " + str(len(sorted_sums)) + " - " + str(sorted_sums))
            
            
        else:
             logger.info(inp_email_id + " -  " + "0 existing allocations found for user - Hence Proceeded")
             returned_total_alloc_status = "Proceed"
                  
    else:
        returned_total_alloc_status  = "_RestCallERROR_ --> " + str(rest_status_code)

    return returned_total_alloc_status

## ---------------------------------------------------------------------------------

def rest_get_Allocation_check_API_Call_Loop(inp_email_id,inp_start_date,inp_end_date, pagination_no):

    created_jql = "project = JRT AND issuetype = 'Daily Allocation' AND status = 'Allocated' AND 'JRT_Resource Email' ~ '" +  str(inp_email_id) + "' and JRT_Date >= '" + str(inp_start_date)[0:10] + "' and  JRT_Date <= '" + str(inp_end_date)[0:10] + "'"
    #created_jql = "project = JRT AND issuetype = 'Daily Allocation' AND status = 'Allocated' AND 'JRT_Resource Email' ~ '" +  str(inp_email_id) + "' and JRT_Date >= '" + str(inp_start_date)[0:10] + "' and  JRT_Date <= '" + str(inp_end_date)[0:10] + "' and 'JRT_Hours per Day[Number]' > " + str(inp_req_hours)
    pagination_startAt = pagination_no
    maxResults = 100

    endpoint = "https://aifel.atlassian.net/rest/api/3/search?"
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA==" # Gobi's Tkn

    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    input_params = {'jql': created_jql,'fields': "customfield_11024,customfield_11030", 'maxResults' : maxResults, 'startAt': pagination_startAt}

    r = requests.get(url=endpoint, headers=header_values,params=input_params, timeout=10)

    if r.status_code == 200:
        response = r.json()
        #record_count = response['total']
                                       
    else:
        response  = "_RestCallERROR_ --> " + str(r.status_code)
        #record_count = "_RestCallERROR_ --> " + str(r.status_code)

    #return r.status_code, response,record_count
    return r.status_code, response


###################################################################

if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "D:\\Python_InputFiles_ForScripts\\RSSD-45_OverAllocation.xlsx"
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 2  # < -- CONFIGURATION
    i_max = 20  # < -- CONFIGURATION
    for i in range(i_start, i_max):
        email_id = xl_sheet.cell(row=i, column=1).value  # < -- Config

        if email_id is not None:
            start_date = xl_sheet.cell(row=i, column=2).value  # < -- Config
            end_date = xl_sheet.cell(row=i, column=3).value  # < -- Config
            req_alloc = xl_sheet.cell(row=i, column=4).value  # < -- Config

            # -->> This is because the user cannot have a single record in system with a allocation of > req_alloc
            #req_alloc = 8 - float(req_alloc)  # This amendment MOVED TO function itself #

            if [start_date, end_date, req_alloc] is not None:
                rest_response = rest_get_Allocation_check_v_1_0(email_id, start_date, end_date, req_alloc)
                xl_sheet.cell(row=i, column=5).value = rest_response
                print(rest_response)
            else:
                error = "One of the input is not valid, Please double check the record -->> " + str(i)
                xl_sheet.cell(row=i, column=5).value = error
                print(error)

    xl_workbook.save(path)

#######################################################################


