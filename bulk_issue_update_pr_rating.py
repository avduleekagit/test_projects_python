import requests
import openpyxl


def rest_PUT_Issue_Edit(ticket_key,reviewer_id, review_eligibility):

    # print(str(reviewer_id) + " --->> within function --->> " + str(review_eligibility))

    if reviewer_id is not None:

        json_data = {
            "fields": {
                "customfield_11271": {
                    "accountId": reviewer_id
                },
                "customfield_11269": {
                    "value": review_eligibility
                }
            }
        }

    else:

        json_data = {
            "fields": {
                "customfield_11271": {
                    "accountId": None
                },
                "customfield_11269": {
                    "value": review_eligibility
                }
            }
        }

    # print(json_data)


    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + str(ticket_key)
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA=="  #Gobi's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.put(url=endpoint, headers=header_values,json=json_data, timeout=10)

    if r.status_code == 204:
        returned_response = "Updated"
    else:
        returned_response = "Error when updating the Issue_" + str(ticket_key) + "_<<RestCallERROR>>_" + str(r.status_code) + "---" + str(r.json())

    return returned_response


if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "D:\\Roshinie_105953_Backup\\ADL\\VSCodeProjects\\script_inputs\\INDO & MY Project List for Evaluation-INDO Finalized.xlsx"
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 35  # < -- CONFIGURATION
    i_max = 44 # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_key = xl_sheet.cell(row=i, column=1).value  # < -- Config

        drop_reviewer_id = xl_sheet.cell(row=i, column=6).value  # < -- Config
        drop_review_eligibility = xl_sheet.cell(row=i, column=5).value  # < -- Config

        if [ticket_key, drop_review_eligibility] is not None:

            # print(str(drop_review_eligibility) + "--->>" + str(drop_reviewer_id))

            update_result = rest_PUT_Issue_Edit(ticket_key,drop_reviewer_id,drop_review_eligibility)
            print (str(ticket_key) + " - - - >> " + str(update_result))
            xl_sheet.cell(row=i, column=8).value = update_result

    xl_workbook.save(path)