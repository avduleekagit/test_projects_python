import requests
import logging
import openpyxl
import json
import math
import logging
import datetime

from Bulk_Allocation_Creation_Functions import validate_record_main_function,direct_allocation_req_class,rest_Post_Issue_Create_for_Direct_Allocation_Req_SD
from Check_Over_Allocation import rest_get_Allocation_check_v_1_0


###################################################################

if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    ### ============================================================
    path = "D:\\Roshinie_105953_Backup\\ADL\\VSCodeProjects\\Bulk Allocations\\RSSD-370.xlsx"
    reporter_account_id = '5d9afd7195a1690dcbbe226f'  #< -- CHANGE THIS   / / / / /  IMPORTANT CONFIGURATION -- IMPORTANT -- >
    i_start = 2  # < -- CONFIGURATION
    i_max = 22  # < -- CONFIGURATION
    ### ============================================================


    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active


    for i in range(i_start, i_max):
        summary = xl_sheet.cell(row=i, column=1).value  # < -- Config
        
        if summary is not None:
            project_code= xl_sheet.cell(row=i, column=2).value  # < -- Config
            project_name= xl_sheet.cell(row=i, column=3).value  # < -- Config
            resour_email= xl_sheet.cell(row=i, column=4).value  # < -- Config
            allo_percent= xl_sheet.cell(row=i, column=5).value  # < -- Config
            all_start= xl_sheet.cell(row=i, column=6).value  # < -- Config
            all_end= xl_sheet.cell(row=i, column=7).value  # < -- Config
            bill_type= xl_sheet.cell(row=i, column=8).value  # < -- Config
            is_over_allo= xl_sheet.cell(row=i, column=9).value  # < -- Config
            special_app_req= xl_sheet.cell(row=i, column=10).value  # < -- Config
            is_app_est_attched = xl_sheet.cell(row=i, column=11).value  # < -- Config
            descrip= xl_sheet.cell(row=i, column=12).value  # < -- Config
            is_extra_resource = xl_sheet.cell(row=i, column=13).value  # < -- Config  18/05/23 Change Req
            summary= str(" [ Bulk Allocation Req ] | OverAlloc= " ) + is_over_allo + str (" | ") + summary  # < -- Ref RSSD - 248 

            #ref_ticket_id = xl_sheet.cell(row=i, column=15).value  # < -- Config
            record_processed_status = xl_sheet.cell(row=i, column=14).value  # < -- Config

            # make sure the date formats
            print(all_start)
            print(all_end)
            all_start=datetime.date.strftime(all_start, "%Y-%m-%d")
            all_end=datetime.date.strftime(all_end, "%Y-%m-%d")

            if [project_code, project_name, resour_email,allo_percent,all_start,all_end,bill_type,is_over_allo,special_app_req,is_app_est_attched] is not None and (allo_percent > 0 and allo_percent <=100) :
                if record_processed_status is None:
                    print("START ------> " + str(resour_email) + " | " + str(all_start) + " to " + str(all_end) + " - validating record .....")
                    valid_status,resour_account_id,resour_allocated_hours = validate_record_main_function(all_start,all_end,project_code,resour_email)            
                    
                    if valid_status == "Valid" and [resour_account_id,resour_allocated_hours] is not None:
                        
                        #req_hours = 8 * (allo_percent/100) ##  - - - - - This needs to be Fixed with the change in rest_get_Allocation_check_v_1_0
                        req_hours = resour_allocated_hours * (allo_percent/100)  ## -- 18/05/23
                        req_hours = round(req_hours,2)
                        print ("Requested hours --> " + str(req_hours))

                        if is_over_allo=="No":
                            print("Executing over allocation check ......")
                            get_allocation_check = rest_get_Allocation_check_v_1_0(resour_email, all_start, all_end, req_hours,resour_allocated_hours)
                            is_over_allo="Bulk"
                        else: # -- Yes scenario
                            get_allocation_check = "Proceed"
                        
                        if get_allocation_check == "Proceed":
                            
                            create_ticket_obj = direct_allocation_req_class()
                            create_ticket_obj.reporter_account_id = reporter_account_id
                            create_ticket_obj.summary=summary
                            create_ticket_obj.project_name_code =project_code
                            create_ticket_obj.project_name = project_name
                            create_ticket_obj.resource_account_id = resour_account_id
                            create_ticket_obj.alloc_percent= allo_percent
                            create_ticket_obj.alloc_start=all_start
                            create_ticket_obj.alloc_end=all_end
                            create_ticket_obj.bill_type=bill_type
                            create_ticket_obj.allocation_check=is_over_allo # -- Manipulation done # - 
                            create_ticket_obj.special_app=special_app_req
                            create_ticket_obj.estimation_attched=is_app_est_attched
                            create_ticket_obj.description=descrip
                            create_ticket_obj.is_extra_resource=is_extra_resource # 18/05/23 Change
                        
                            # -- Finally create Ticket
                            print("END --------->  Creating ticket ")
                            record_processed_status = rest_Post_Issue_Create_for_Direct_Allocation_Req_SD(create_ticket_obj)
                            #record_processed_status ="Validation_Ok"
                        else:
                            record_processed_status = get_allocation_check
                    
                    else:
                        record_processed_status = "Record validation failure --> " + str(valid_status)
                    
            else:
                record_processed_status = "Input validation error : One of the input is either invalid or unavailable, Please double check the record !"


            ## Update Excel ----
            xl_sheet.cell(row=i, column=14).value = record_processed_status

    xl_workbook.save(path)

#######################################################################