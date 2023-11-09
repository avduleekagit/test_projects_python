import requests
import logging
import openpyxl
from datetime import datetime

from Check_Over_Allocation import rest_get_Allocation_check_v_1_0
from Child_issue_create import rest_get_Parent_Details,rest_Post_Issue_Create
from Working_Day_checker import get_working_days
### For account status check CR ---###### --------------
from check_user_status import rest_get_UserStatus_Check,rest_get_user_validations_jql_call
### For account status check CR ---###### --------------

############################################################
#rest_response = rest_get_Allocation_check("parthiban.chandrasegar@axiatadigitallabs.com", "2023-01-01", "2023-03-31", "5")
#print(rest_response)

#get_parent_ticket_details = rest_get_Parent_Details('JRT-1351')


###====================================================================
def Process_Allocation(parent_ticket):

    return_message = parent_ticket #just to assign a default value to return_message and append accoridingly

    get_parent_ticket_details = rest_get_Parent_Details(parent_ticket)

    email = get_parent_ticket_details.resource_email
    start_date = get_parent_ticket_details.alloc_start
    ##end_date = get_parent_ticket_details.alloc_end  ## Requested by Nuwanee
    end_date = get_parent_ticket_details.confirmed_end
    hours = get_parent_ticket_details.hou_per_day
    is_over_allocation = get_parent_ticket_details.is_over_allocation # 23/06/2023 To handle JRT_Over Allocation records
     
    ### For User Jira Account status check CR ---requested by nuwanee  ###### --------------
    user_account_id = get_parent_ticket_details.resource_name
    user_account_status = rest_get_UserStatus_Check (user_account_id)
    if user_account_status != "Proceed":
        return_message = user_account_status
        return return_message
    
    ### For User Jira Account status check CR -----------------------------###### 

    ## For  Resource card status check CR - requested by nuwanee --######### -----------------
    jql = "project = JRT AND issuetype = 'RM Resource Creation' AND status = Active AND 'JRT_Resource Email[Short text]' ~ " + "'" + str(email) + "'"
    req_fields= "customfield_10968,customfield_10978"
    get_resrce_card_detail_status_code, get_resrce_card_detail_response_json = rest_get_user_validations_jql_call (jql,req_fields)

    if get_resrce_card_detail_status_code == 200:
             if get_resrce_card_detail_response_json['total'] != 1:
                  return_message = "Error : No Resource card available"
                  return return_message
             
             max_hours_allocation = get_resrce_card_detail_response_json['issues'][0]['fields']['customfield_10978']  ## 18/05/23 CR Req - get maximum working hours from resource card
             print ("max_hours_allocation for user is --> " + str(max_hours_allocation))
    else:
        return_message = "Error REST Calling resource_card_details_check ! " + str(get_resrce_card_detail_response_json)
        return return_message

    # For  Resource card status check --------------------------------######### -----------------

    ##########################################

    if is_over_allocation == "Yes":   # --- 23/06/2023 Fix for Over_allocation "Yes" scenario
        get_allocation_check = "Proceed"
        print ("# --------------- Over Allocation Check Skipped ..... !!! ")
    else:
        get_allocation_check = rest_get_Allocation_check_v_1_0(email, start_date, end_date, hours,max_hours_allocation) # Fix Done with v1.0 19/03/2023 # 18/05/23 CR get max_hours from resource card
    
    if get_allocation_check == "Proceed":
        print(str(parent_ticket) + " |-->>> " + str(email) + " - " + str(start_date) + " - " + str(end_date) + " - " + str(hours) + " --> Proceed ")

        working_days = get_working_days(start_date, end_date)

        task_sq = 0
        for item in working_days:
            sub_task_id = rest_Post_Issue_Create(get_parent_ticket_details, task_sq, item)
            terminal_output = str(task_sq) + "-->" + str(item) + "-->" + str(sub_task_id)
            print(terminal_output)
            
            return_message = str(return_message) + " | " + str(terminal_output)
            task_sq = task_sq + 1

            ## ----  Fix for if 403 error fired during executions for 1 or more issues
            if sub_task_id[0:14] =="Issue Creating":
                return return_message        
            ### -----------------------------------------------------------------------  

        return_message = str(return_message) + " | _Process_Completed_"  # -- To check whther process completed fully


    else:
        return_message = return_message + " | " + str(email) + " is Over Allocated --> " + str(get_allocation_check)
        print(return_message)
    
    return return_message

###====================================================================

if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "C:\\Users\\Gobiananth_105358.ADL\\Downloads\\RM Approved Request Validation_20230217_Gobi List.xlsx" #< -- CONFIGURATION
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 2  # < -- CONFIGURATION
    i_max = 50  # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_id = xl_sheet.cell(row=i, column=1).value  # < -- Config
        ticket_id_status = xl_sheet.cell(row=i, column=9).value  # < -- Config

        if ticket_id is not None: 
            if ticket_id_status != "Processed" :
                rest_response = Process_Allocation(ticket_id)
                xl_sheet.cell(row=i, column=10).value = rest_response
                xl_sheet.cell(row=i, column=9).value = "Processed"
            else: 
                print ("Record already processed , Hence Skipped")
             
        xl_workbook.save(path)
    
    print("Process Completed" + " ==> " + str(datetime.now()))

