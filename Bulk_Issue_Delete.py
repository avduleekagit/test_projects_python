import requests
import logging
import openpyxl
from datetime import datetime

class jira_ticket_class:
  def __init__(self):
    self.issue_id =""
    self.return_response_code=""

def rest_get_ticket_details(ticket_key):

    jira_ticket_obj = jira_ticket_class()
    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + ticket_key
    token = "Basic ZHVsZWVrYS5yb3NoaW5pZUBheGlhdGFkaWdpdGFsbGFicy5jb206SVJPSW1kV3lHeE5QRHFBdVJmQlZCMkQ3"  # Roshinie's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    #input_params = {'fields': "customfield_10987,customfield_11011"}
    #r = requests.get(url=endpoint, headers=header_values,params=input_params, timeout=10)
    r = requests.get(url=endpoint, headers=header_values, timeout=10)
    
    if r.status_code == 200:
        response = r.json()
        jira_ticket_obj.issue_id= response['id']
        jira_ticket_obj.return_response_code = r.status_code

    else:
        jira_ticket_obj.return_response_code = "rest_get_ticket_details_RestCallERROR_" + str(r.status_code)

    return jira_ticket_obj

def rest_get_tempo_worklogs(ticket_id):
   endpoint = "https://api.tempo.io/4/worklogs?issueId=" + str(ticket_id)
   token="Bearer tukFXLvU1wAZuURRdGVSjo3u2JljrU-us"
   header_values = {'Authorization':token}
   r = requests.get(url=endpoint, headers=header_values, timeout=10)
   
   worklog_count = ""

   if r.status_code == 200:
        response = r.json()
        worklog_count= response['metadata']['count']
        return_response_code = r.status_code
   else:
        return_response_code = "rest_get_tempo_worklogs_RestCallERROR_" + str(r.status_code)

   return return_response_code,worklog_count

def rest_DEL_ticket(ticket_key):
    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + ticket_key
    token = "Basic ZHVsZWVrYS5yb3NoaW5pZUBheGlhdGFkaWdpdGFsbGFicy5jb206SVJPSW1kV3lHeE5QRHFBdVJmQlZCMkQ3"  # Roshinie's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.delete(url=endpoint, headers=header_values, timeout=10)

    if r.status_code == 204:
        return_response ="Deleted"
    else:
        return_response = "rest_DEL_ticket_RestCallERROR_" + str(r.status_code) + str(r.json()['errorMessages'])
    
    return return_response

if __name__ == "__main__":
 
    path = "D:\\Roshinie_105953_Backup\\ADL\\VSCodeProjects\\script_inputs\\AJS-7181_log.xlsx" #< -- CONFIGURATION
    #D:\Roshinie_105953_Backup\ADL\VSCodeProjects\script_inputs
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 1  # < -- CONFIGURATION
    i_max = 2  # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_key = xl_sheet.cell(row=i, column=1).value  # < -- Config
        processed_status = xl_sheet.cell(row=i, column=3).value  # < -- Config

        if ticket_key is not None: 
            if processed_status is None :
                print (ticket_key)
                get_ticket_obj = rest_get_ticket_details (ticket_key)
                
                if get_ticket_obj.return_response_code == 200:
                   get_worklog_res_code, get_worklog_count = rest_get_tempo_worklogs (get_ticket_obj.issue_id)
              
                   if get_worklog_res_code == 200:
                       if get_worklog_count == 0:
                            return_response = rest_DEL_ticket(ticket_key)
                            # return_response = "Can be Deleted"
                       else:
                           return_response = "Issue Delete Aborted ! " + str(get_worklog_count) + " " + "worklogs found"
                   else:
                       return_response = get_worklog_res_code        
                      
                else:
                    return_response = get_ticket_obj.return_response_code

                print (return_response)
                xl_sheet.cell(row=i, column=3).value = return_response
            else: 
                print ("Record already processed , Hence Skipped")
             
        xl_workbook.save(path)
    
    print("Process Completed" + " ==> " + str(datetime.now()))