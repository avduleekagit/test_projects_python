import requests
import openpyxl


def rest_PUT_Issue_Edit(ticket_key,bu_name):

    json_data = {
    "fields": {
        "customfield_11241": {
            "value": bu_name
                },
            }
    }

    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + str(ticket_key)
    token = "Basic ZHVsZWVrYS5yb3NoaW5pZUBheGlhdGFkaWdpdGFsbGFicy5jb206SVJPSW1kV3lHeE5QRHFBdVJmQlZCMkQ3"  #Roshinie's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.put(url=endpoint, headers=header_values,json=json_data, timeout=10)

    if r.status_code == 204:
        returned_response = "Updated"
    else:
        returned_response = "Erron when updating the Issue_" + str(ticket_key) + "_<<RestCallERROR>>_" + str(r.status_code) + "---" + str(r.json())

    return returned_response


if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "C:\\Users\\Roshinie_105953\\Downloads\\BU_update_Execution_3.xlsx"
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 101  # < -- CONFIGURATION
    i_max = 130 # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_key = xl_sheet.cell(row=i, column=1).value  # < -- Config
        exis_bu_name = xl_sheet.cell(row=i, column=2).value  # < -- Config

        drop_bu_name = xl_sheet.cell(row=i, column=3).value  # < -- Config
        

        if [ticket_key, drop_bu_name ] is not None:

            update_result = rest_PUT_Issue_Edit(ticket_key,drop_bu_name)
            print (str(ticket_key) + " - - - >> " + str(update_result))
            xl_sheet.cell(row=i, column=4).value = update_result

    xl_workbook.save(path)