import requests
import logging
import math
import datetime
from Jira_RM_Tool import Process_Allocation
from Child_issue_create import rest_PUT_Issue_Edit_for_transChange,rest_Post_Issue_Transisiton,rest_get_Parent_Details
from Change_Allocation_Executor import Process_change_request
import openpyxl


def rest_get__pending_que_items(max_records):

    jql="project = JRT AND issuetype = 'Resource Allocation' AND status = 'RM Approved' ORDER BY updated ASC"
    endpoint = "https://aifel.atlassian.net/rest/api/3/search?"
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA==" # Gobi's Tkn

    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    input_params = {'jql': jql,'fields': 'key,customfield_10021', 'maxResults' : max_records }

    r = requests.get(url=endpoint, headers=header_values,params=input_params, timeout=10)

    if r.status_code == 200:
        response = r.json()
                                       
    else:
        response  = "Error fetching Pending Que Items_RestCallERROR_ --> " + str(r.status_code)

    return r.status_code, response


## ---------########################################################################
excluded_tikets =[]
max_record_fetch_count=10

file_name="Processed_Allocations_Log_" + str (datetime.date.strftime(datetime.datetime.now(), "%Y-%m-%d__%H-%M-%S")) + ".xlsx"
path = "D:\\RM Scheduler Logs\\" + str(file_name) #< -- CONFIGURATION
xl_workbook = openpyxl.Workbook()
xl_sheet = xl_workbook.active

## - - - - -##########################################################################


if __name__ == "__main__":

    result_code,que_items_result = rest_get__pending_que_items(max_record_fetch_count)
    execuded_ticket_count =0 # --  variable initial assignment 
    ticket_cursor_id =1 # --  variable initial assignment 
    post_ticket_update_status = "" # --  variable initial assignment 
    post_ticket_transition_status = "" # --  variable initial assignment 

    if result_code == 200:
        if que_items_result['total'] > 0:

            for issue in que_items_result['issues']:
                ticket_id = issue['key']
                ticket_type = issue['fields']['customfield_10021']['requestType']['name']
                
                if ticket_id not in excluded_tikets:
                    ticket_cursor_id = ticket_cursor_id + 1
                    print(ticket_id)
                    print(ticket_type)
                    
                    if ticket_type == "Direct Resource Allocation Request":
                        processed_status = Process_Allocation(ticket_id)
                        if processed_status[-19:] == "_Process_Completed_":
                            post_ticket_update_status = rest_PUT_Issue_Edit_for_transChange(ticket_id,'Script')
                            post_ticket_transition_status = rest_Post_Issue_Transisiton(ticket_id,131)
                            return_response = str(processed_status) + " | " + str(post_ticket_update_status) + " | " + str(post_ticket_transition_status)
                        else:
                            return_response = processed_status
                            break
                    
                    if ticket_type == "Change Allocation Request":     
                        get_change_ticket_obj = rest_get_Parent_Details(ticket_id)
                        get_orig_ticket_obj = rest_get_Parent_Details(get_change_ticket_obj.orig_tkt_id_for_change)
                        processed_status = Process_change_request(get_change_ticket_obj,get_orig_ticket_obj)
                        if processed_status[-19:] == "_Process_Completed_":
                            post_ticket_update_status = rest_PUT_Issue_Edit_for_transChange(ticket_id,'Script')
                            post_ticket_transition_status = rest_Post_Issue_Transisiton(ticket_id,131)
                            return_response = str(processed_status) + " | " + str(post_ticket_update_status) + " | " + str(post_ticket_transition_status)
                        else:
                            return_response = processed_status
                            break
                    
                    ##### -- Update Excel  ----##
                    xl_sheet.cell(row=ticket_cursor_id, column=1).value = ticket_id
                    xl_sheet.cell(row=ticket_cursor_id, column=2).value = ticket_type
                    xl_sheet.cell(row=ticket_cursor_id, column=3).value = processed_status[-19:]
                    xl_sheet.cell(row=ticket_cursor_id, column=4).value = post_ticket_update_status
                    xl_sheet.cell(row=ticket_cursor_id, column=5).value = post_ticket_transition_status
                    xl_sheet.cell(row=ticket_cursor_id, column=6).value = str (datetime.date.strftime(datetime.datetime.now(), "%Y-%m-%d__%H-%M-%S"))    
                    xl_sheet.cell(row=ticket_cursor_id, column=7).value = return_response
                    xl_workbook.save(path)

                else:
                    execuded_ticket_count = execuded_ticket_count + 1
                    if  execuded_ticket_count == len (excluded_tikets):
                            return_response = "No Records found to process after exclusions"

        else:
            return_response = "No Records found to process"
    else:
        return_response = que_items_result

    print(return_response)
    print("Process Completed at" + " ==> " + str(datetime.datetime.now()))
