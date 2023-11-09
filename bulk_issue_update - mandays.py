import requests
import openpyxl
import json

def rest_Post_Issue_AddComment(ticket_key):

    false = False
    json_data = json.dumps({
        "body": "Planned Man Days value has been changed referring a bulk update request through RSSD-305",
        "public": false      
    })

    endpoint = "https://aifel.atlassian.net/rest/servicedeskapi/request/" + str(ticket_key) + "/comment"
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA=="  #Gobi's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.post(url=endpoint, headers=header_values,data=json_data, timeout=10)

    if r.status_code == 201:
        returned_response = "CommentAdded"
    else:
        returned_response = "Error when adding comment_" + str(ticket_key) + "_<<RestCallERROR>>_" + str(r.status_code) + "---" + str(r.json())

    return returned_response


def rest_PUT_Issue_Edit(ticket_key,man_days):

    json_data = {
    "fields": {
        "customfield_11236": man_days
            }
    }

    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + str(ticket_key)
    token = "Basic bmFkZXNhcGlsbGFpLmdvYmlhbmFudGhAYXhpYXRhZGlnaXRhbGxhYnMuY29tOmpJWXQ0cnhPeFdtMU81VWdQZ2lkODM2MA=="  #Gobi's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.put(url=endpoint, headers=header_values,json=json_data, timeout=10)

    if r.status_code == 204:
        returned_response = "Updated"
    else:
        returned_response = "Error when updating the Issue_" + str(ticket_key) + "_<<RestCallERROR>>_" + str(r.status_code) + "---" + str(r.json())

    return returned_response


if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "D:\\Roshinie_105953_Backup\\ADL\\VSCodeProjects\\script_inputs\\Proposed Mandays_JIRA Master Data.xlsx"
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 131  # < -- CONFIGURATION
    i_max = 153 # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_key = xl_sheet.cell(row=i, column=1).value  # < -- Config

        new_man_days = xl_sheet.cell(row=i, column=4).value  # < -- Config

        if [ticket_key, new_man_days] is not None:

            update_result = rest_PUT_Issue_Edit(ticket_key,new_man_days)
            update_comment_result = rest_Post_Issue_AddComment(ticket_key)
            print (str(ticket_key) + " - - - >> " + str(update_result) + " - - - >> " + str(update_comment_result))
            xl_sheet.cell(row=i, column=8).value = update_result
            xl_sheet.cell(row=i, column=9).value = update_comment_result

    xl_workbook.save(path)