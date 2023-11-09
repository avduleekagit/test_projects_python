import requests
import openpyxl


def rest_PUT_Issue_Edit(ticket_key):

    json_data = {
    "fields": {
        "customfield_11145": {
            "value": "N/A"
                },
            }
    }

    endpoint = "https://aifel.atlassian.net/rest/api/3/issue/" + str(ticket_key)
    token = "Basic amlyYUFkbWluc0BheGlhdGFkaWdpdGFsbGFicy5jb206QVRBVFQzeEZmR0YwRUdiekNIQW1iSHZfUXc2VzNiNXdidjdISjlxeEhHSWlmbUFVdFNYZWlYcl9CZ3pMSXNOX3hScUc4dUtjdy00ajk0UG1YNFVhbTNtZTBZdnVPa1FOSGFyMnFCQ1NUQUFLV3FVNTJKQ2hDOUlud2stQ2dXMTlzNUZxM0g5Ujd2NUQtc1U1YnRfd3NoR29JODduN2pJeEhxTkVYbGpaOGFTcE0tQ3NHUHpYbjJBPTY5QjM4MEY0"  #Jira Admin's Token
    
    header_values = {'Content-Type': "application/json",'Accept': "application/json",'Authorization':token}
    r = requests.put(url=endpoint, headers=header_values,json=json_data, timeout=10)

    if r.status_code == 204:
        returned_response = "Updated"
    else:
        returned_response = "Erron when updating the Issue_" + str(ticket_key) + "_<<RestCallERROR>>_" + str(r.status_code) + "---" + str(r.json())

    return returned_response


if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "D:\\Roshinie_105953_Backup\\ADL\\JIRA\\RM Tool Related Docs\\Bulk Update\\Project Card Export - 24102023.xlsx"

    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 251  # < -- CONFIGURATION
    i_max = 292 # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_key = xl_sheet.cell(row=i, column=1).value  # < -- Config
        exis_secondary_project_type = xl_sheet.cell(row=i, column=2).value  # < -- Config

        updated_secondary_project_type = xl_sheet.cell(row=i, column=3).value  # < -- Config
        

        if [ticket_key ] is not None:

            update_result = rest_PUT_Issue_Edit(ticket_key)
            print (str(ticket_key) + " - - - >> " + str(update_result))
            xl_sheet.cell(row=i, column=4).value = update_result
            xl_sheet.cell(row=i, column=3).value = "N/A"
    
    xl_workbook.save(path)