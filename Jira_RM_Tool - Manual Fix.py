import requests
import logging
import openpyxl
from datetime import datetime

from Check_Over_Allocation import rest_get_Allocation_check
from Child_issue_create import rest_get_Parent_Details,rest_Post_Issue_Create
from Working_Day_checker import get_working_days

############################################################
#rest_response = rest_get_Allocation_check("parthiban.chandrasegar@axiatadigitallabs.com", "2023-01-01", "2023-03-31", "5")
#print(rest_response)

#get_parent_ticket_details = rest_get_Parent_Details('JRT-1351')


###====================================================================
def Process_Allocation(parent_ticket,record_kicker):

    return_message = parent_ticket #just to assign a default value to return_message and append accoridingly

    get_parent_ticket_details = rest_get_Parent_Details(parent_ticket)

    email = get_parent_ticket_details.resource_email
    start_date = get_parent_ticket_details.alloc_start
    ##end_date = get_parent_ticket_details.alloc_end
    end_date = get_parent_ticket_details.confirmed_end
    hours = get_parent_ticket_details.hou_per_day

    #get_allocation_check = rest_get_Allocation_check(email, start_date, end_date, hours)
    #if get_allocation_check == "Proceed":
    print(str(parent_ticket) + " |-->>> " + str(email) + " - " + str(start_date) + " - " + str(end_date) + " - " + str(hours) + " --> Allocation Check Skipped due to failed record processing . . . . ")

    working_days = get_working_days(start_date, end_date)

    task_sq = 0
    for item in working_days:
            if task_sq > record_kicker: ## Input the Record id where the execution stopped during the run ##
                sub_task_id = rest_Post_Issue_Create(get_parent_ticket_details, task_sq, item)
                terminal_output = str(task_sq) + "-->" + str(item) + "-->" + str(sub_task_id)
                print(terminal_output)
            
                return_message = str(return_message) + " | " + str(terminal_output)           
            else: 
                print(str(task_sq) + "-->" + str(item) + " --> sequence already processed")
            task_sq = task_sq + 1
    #else:
        #return_message = return_message + " | " + str(email) + " is Over Allocated --> " + str(get_allocation_check)
        #print(return_message)
    
    return return_message

###====================================================================

if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "D:\\Roshinie_105953_Backup\\ADL\\VSCodeProjects\\RM Approved Request Validation_20230927_Failed List_executed.xlsx" #< -- CONFIGURATION
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 2  # < -- CONFIGURATION
    i_max = 4  # < -- CONFIGURATION
    for i in range(i_start, i_max):
        ticket_id = xl_sheet.cell(row=i, column=1).value  # < -- Config
        ticket_id_status = xl_sheet.cell(row=i, column=9).value  # < -- Config
        record_stopped_at = xl_sheet.cell(row=i, column=2).value  # < -- Config

        if ticket_id is not None:
            if ticket_id_status != "Processed" :
                rest_response = Process_Allocation(ticket_id,record_stopped_at)
                xl_sheet.cell(row=i, column=10).value = rest_response
                xl_sheet.cell(row=i, column=9).value = "Processed"
            else: 
                print ("Record already processed , Hence Skipped")
             
        xl_workbook.save(path)
    
    print("Process Completed" + " ==> " + str(datetime.now()))

