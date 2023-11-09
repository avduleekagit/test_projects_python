import requests
import openpyxl
from Child_issue_create import rest_get_Parent_Details,rest_get__jql_result_Loop,rest_get_full_jql_result,rest_Post_Issue_Transisiton,rest_Post_Issue_AddComment


if __name__ == "__main__":
   # stuff only to run when not called via 'import' here

    path = "D:\\Roshinie_105953_Backup\\ADL\\VSCodeProjects\\Project Card Planning Lines\\Planning_lines - RSSD-438.xlsx"
    xl_workbook = openpyxl.load_workbook(path)
    xl_sheet = xl_workbook.active

    i_start = 1  # < -- CONFIGURATION
    i_max = 14 # < -- CONFIGURATION
    for i in range(i_start, i_max):
        parent_ticket_id = xl_sheet.cell(row=i, column=1).value  # < -- Config

        if parent_ticket_id is not None:
            parent_ticket_details = rest_get_Parent_Details(parent_ticket_id)
            alloc_start_date=parent_ticket_details.alloc_start
            alloc_confirm_end_date=parent_ticket_details.confirmed_end

            jql = "project=JRT and issuetype = 'Daily Allocation' and status = Allocated and 'JRT_Demand Reference ID[Short text]' ~ " + parent_ticket_id + " and ( 'JRT_Date[Date]' < " +  alloc_start_date + " or 'JRT_Date[Date]' > " + alloc_confirm_end_date + " )"
            req_fields= "key"    

            ## ----
            get_daily_records_to_cancel= rest_get_full_jql_result(jql,req_fields)
            ## ----

            # -- Cancel Issues + Add Comment
            if len(get_daily_records_to_cancel) > 0:
                returned_response = "Error ! " + str(parent_ticket_id) + " - " + str (len(get_daily_records_to_cancel)) + " out of range records found ! ... Cancelling daily records --->> "
                print (returned_response)
                issue_comment="This task was cancelled referring the fix - RSSD-438"
                
                for cancel_issue_key in get_daily_records_to_cancel:
                  print("Cancelling . . . " + str(cancel_issue_key))
                  rest_issue_cancel_response = rest_Post_Issue_Transisiton(cancel_issue_key,21)
                  rest_issue_addComment_response = rest_Post_Issue_AddComment(cancel_issue_key,issue_comment)

                  if (rest_issue_cancel_response != "Transitioned") or (rest_issue_addComment_response != "CommentAdded"):
                        returned_response = str(returned_response) + " | " + str(rest_issue_cancel_response) +  " | " + str(rest_issue_addComment_response)
                        print (str(cancel_issue_key) + " - - - - -" + " Execution error - cancel/comment_update,Hence process aborted -------------!")
                        break
                  else:
                        returned_response = str(returned_response) + " | " + str(cancel_issue_key)
            
            else:
                returned_response = "Ok"
                print (returned_response)
            
            xl_sheet.cell(row=i, column=2).value = returned_response
            xl_sheet.cell(row=i, column=3).value = len (get_daily_records_to_cancel)
            xl_sheet.cell(row=i, column=4).value = str(get_daily_records_to_cancel)

    xl_workbook.save(path)